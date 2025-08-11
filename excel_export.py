# excel_export.py — IvyRecon branded Excel exporter
from __future__ import annotations
from typing import Optional, Tuple
from io import BytesIO
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.worksheet.worksheet import Worksheet

TEAL = "18CCAA"
NAVY = "2F455C"
HEADER_FILL = PatternFill("solid", fgColor=TEAL)
HEADER_FONT = Font(color="0F2A37", bold=True)
BORDER_THIN = Border(
    left=Side(style="thin", color="DDDDDD"),
    right=Side(style="thin", color="DDDDDD"),
    top=Side(style="thin", color="DDDDDD"),
    bottom=Side(style="thin", color="DDDDDD"),
)

def _autofit(ws: Worksheet):
    widths = {}
    for row in ws.iter_rows():
        for cell in row:
            val = "" if cell.value is None else str(cell.value)
            # modest width calc
            w = min(max(len(val) + 2, 6), 50)
            if cell.column_letter not in widths or w > widths[cell.column_letter]:
                widths[cell.column_letter] = w
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

def _style_sheet(ws: Worksheet, title: Optional[str] = None):
    if title:
        ws["A1"].value = title
        ws["A1"].font = Font(color=NAVY, bold=True, size=12)
    # find header row (first data row)
    hdr_row = 2 if title else 1
    max_col = ws.max_column
    for c in range(1, max_col + 1):
        cell = ws.cell(row=hdr_row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = BORDER_THIN
        cell.alignment = Alignment(vertical="center")
    # borders & align
    for r in ws.iter_rows(min_row=hdr_row + 1, max_row=ws.max_row, max_col=max_col):
        for cell in r:
            cell.border = BORDER_THIN
            if isinstance(cell.value, (int, float)):
                cell.alignment = Alignment(horizontal="right", vertical="center")
            else:
                cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = ws["A" + str(hdr_row + 1)]
    _autofit(ws)

def _write_df(ws: Worksheet, df: pd.DataFrame, title: Optional[str] = None):
    # optional title row
    r0 = 1
    if title:
        ws.append([title])
        r0 += 1
    # dataframe rows
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)
    _style_sheet(ws, title)

def export_errors_multitab(errors_df: pd.DataFrame, summary_df: pd.DataFrame,
                           group_name: str = "", period: str = "") -> bytes:
    wb = Workbook()
    wb.remove(wb.active)

    # Summary
    ws_sum = wb.create_sheet("Summary")
    _write_df(ws_sum, summary_df if summary_df is not None else pd.DataFrame(),
              title=f"Summary — {group_name or ''} {period or ''}".strip())

    # All Errors
    ws_all = wb.create_sheet("All Errors")
    _write_df(ws_all, errors_df if errors_df is not None else pd.DataFrame(),
              title=f"All Errors — {group_name or ''} {period or ''}".strip())

    # Per-type sheets
    if errors_df is not None and not errors_df.empty and "Error Type" in errors_df.columns:
        for etype, chunk in errors_df.groupby("Error Type", sort=False):
            safe = "".join(ch for ch in str(etype) if ch.isalnum() or ch in (" ", "_", "-"))[:28]
            ws = wb.create_sheet(safe or "Errors")
            _write_df(ws, chunk.reset_index(drop=True), title=str(etype))

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.getvalue()

